"""
    criado em 19/09/2023
    autor: Eduardo Santos
    ultima modificacao: 21/09/2023
"""
import os # Biblioteca de interação com Sistema Operacional
import pymysql # Biblioteca para fazer interacao com BD
import openpyxl # Biblioteca para ler planilha xlsx
import prompt_toolkit # Biblioteca para usar autocomplete nos input
import prompt_toolkit.completion

# Configurar o completador de caminhos para o diretório atual
completer = prompt_toolkit.completion.PathCompleter()

# Conexao com BD
banco = pymysql.connect(    
    host="localhost",
    user="root",
    passwd="root",
    database="login"
)

# Definir Cursor
cursor = banco.cursor()

pasta = './' # Torna a pasta atual o diretório de trabalho
arqs = os.listdir(pasta) # Lista os arquivos locais

planilhas = [] # Array auxiliar
format_visual = '-' * 50 # Quantidade de '-' que serão impressos

# Verifica se há alguma planilha '.xlsx' no diretório local
for(arq) in arqs: 
    if os.path.splitext(arq)[1] == ".xlsx":
        planilhas.append(arq)

if len(planilhas) <=0:
    print('Nenhum arquivo ".xlsx" foi encontrado!')
else:
    print(format_visual)
    print("Os seguintes arquivos foram encontrados:")
    print(format_visual)
    for i in planilhas:
        print(i)    
    print(format_visual)      
    
# Aguarda o usuário inserir a planilha
print("Insira o arquivo a ser trabalhado: ")
arq=prompt_toolkit.prompt('',completer=completer)

print("\nCarregando arquivo...\n")

# Carregando nosso workbook (Imagine ele como um livro físico para editar as planilhas)
wb = openpyxl.load_workbook(filename=arq) # Abrindo o Arquivo.xlsx

if len(wb.sheetnames) < 1:
    print("Nenhuma planilha foi encontrada, deseja criar uma? (Y/N)")
    answ = input()
    if answ == ('Y' or 'y'):
        print("\nInsira o nome da planilha: ")
        plan_nome = input()
        wb.create_sheet(plan_nome)
        print("\n Criado com sucesso!")
    elif answ == ('N' or 'n'):
        print("\nNenhuma planilha foi criada!")
else:
    print("Selecione uma Planilha para trabalhar:")
    print(format_visual)
    planilhas.clear()
    for i in wb.sheetnames:
        print(i)
        planilhas.append(i)
    print(format_visual)

ws = wb[prompt_toolkit.prompt('',completer=prompt_toolkit.completion.WordCompleter(planilhas))]
print(format_visual)
print('\nA planilha: "{}" foi selecionada.'.format(i))
print(format_visual)

linhas = ws.max_row
colunas = ws.max_column



print('A Planilha contém: "{}" linhas e "{}" colunas.'.format(linhas,colunas))


# sheet = wb['dados_01'] # Selecionando o nome da planilha

# Array auxiliar que vai guardar os dados para serem registrados no BD
arr=[]

# Percorre a planilha
# for x in range(2,32): # Percorre as linhas
#     for y in range (2,6): # Percorre as colunas
#         arr.append(sheet.cell(row=x,column=y).value)

#     # Definir os valores que serão registrados no banco
#     registro = """
#         insert into usuario(nome,sobrenome,email,senha)
#             values('{}','{}','{}','{}');
#         """.format(arr[0],arr[1],arr[2],int(arr[3]))

#     # Realiza o registro
#     cursor.execute(registro)
#     arr.clear()

#Fechar conexoes
cursor.close()    
banco.close()

print(format_visual,"\nFim do Programa.")
'''
 Fim do codigo até 19/09/2023
 modificado por: Eduardo Santos
'''